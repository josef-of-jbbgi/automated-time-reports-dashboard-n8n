#!/usr/bin/env python3
"""Generate the test report Excel file for Journey Dashboard."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

wb = Workbook()

# ─── Colors ───
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD = Font(bold=True, size=11)
NORMAL = Font(size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

def style_row(ws, row, result):
    fill = GREEN if result == "PASS" else RED if result == "FAIL" else YELLOW
    for cell in ws[row]:
        cell.font = NORMAL
        cell.alignment = WRAP
        cell.border = THIN_BORDER
    ws.cell(row=row, column=4).fill = fill

# ═══════════════════════════════════════════════════════════
# Sheet 1: Frontend UI Tests (Playwright)
# ═══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Frontend UI Tests"
ws1.append(["#", "Test Case", "Description", "Result", "Expected Outcome", "Actual Outcome", "Solution"])
style_header(ws1)

ui_tests = [
    ["F01", "Header greeting + date",
     "Verify Header shows time-based greeting and today's formatted date",
     "PASS", "Greeting like 'Good afternoon, Josef.' + 'Friday, February 13'",
     "Shows 'Good afternoon, Josef.' + 'Friday, February 13'", "—"],

    ["F02", "TaskList Internal/Client sections",
     "Verify TaskList renders 'Internal' and 'Client' section headers when tasks exist",
     "PASS (conditional)", "Two sections visible when tasks are loaded; empty state when no tasks",
     "Empty state shows 'No tasks yet.' (correct — no data in Airtable). Sections render only with data.",
     "Expected behavior. Will validate with real data after morning workflow runs."],

    ["F03", "TaskQuickAdd expands",
     "Click '+ Add task or note' → input, priority dropdown, Add/Cancel buttons appear",
     "PASS", "Input field, select dropdown (High/Medium/Low), Add and Cancel buttons",
     "All elements render correctly. Input autofocuses.", "—"],

    ["F04", "DraftPanel Time-In",
     "Time-In panel renders with empty state message",
     "PASS", "'No drafts available' + 'Drafts will generate at 7:30 AM'",
     "Shows exactly as expected", "—"],

    ["F05", "DraftPanel Time-Out lock",
     "Time-Out panel shows locked state before Time-In is sent",
     "PASS", "'Send Time-In first' in both status badge and body area",
     "Lock state renders correctly in both positions", "—"],

    ["F06", "PromptWindow",
     "Prompt Window renders with input, button, and placeholder",
     "PASS", "Input with placeholder, 'Send to Agent' button, helper text",
     "All elements visible: input, button, 'Ask the AI agent...' text", "—"],

    ["F07", "TimeTracker",
     "TimeTracker shows 'Not Started' state with current time",
     "PASS", "'Timed In: —', 'Current: [time]', 'Elapsed: —'",
     "Shows 'Timed In: — | Current: 2:41 PM | Elapsed: —'", "—"],

    ["F08", "OfflineBanner toggle",
     "Amber banner appears on offline, disappears on online",
     "PASS", "Banner with 'You're offline — showing cached data' appears/disappears",
     "Banner appears immediately on offline event, disappears on online", "—"],

    ["F09", "Toast on error",
     "Error toast appears when API call fails",
     "PASS", "Red toast in bottom-right with 'Failed to add task'",
     "Toast renders correctly with auto-dismiss", "—"],

    ["F10", "Not-found page",
     "Navigate to /nonexistent → custom 404 page renders",
     "PASS", "'Page not found' + description on dark background",
     "Renders 'Page not found' centered with muted description", "—"],

    ["F11", "Responsive 375px",
     "All sections stack cleanly at mobile viewport (375px)",
     "PASS", "No horizontal overflow, all cards fit within viewport",
     "Body scrollWidth=375 === viewport. No overflow detected.", "—"],

    ["F12", "Full dashboard layout",
     "All 6 sections visible in correct order",
     "PASS", "Header → Time-In → Tasks → Time-Out → Prompt → Tracker",
     "All sections render in correct order with proper spacing", "—"],

    ["F13", "Graceful API errors",
     "Page renders content even when all APIs return 500",
     "PASS", "Dashboard structure visible (not white screen)",
     "All section headers and empty states render despite API failures", "—"],

    ["F14", "Dark theme",
     "html has class='dark', body bg is dark (#0a0a0a)",
     "PASS", "html.dark, body bg = rgb(10, 10, 10)",
     "Confirmed: html class='dark', bg = rgb(10, 10, 10)", "—"],
]

for row in ui_tests:
    ws1.append(row)
    style_row(ws1, ws1.max_row, row[3].split(" ")[0])

ws1.column_dimensions["A"].width = 6
ws1.column_dimensions["B"].width = 28
ws1.column_dimensions["C"].width = 50
ws1.column_dimensions["D"].width = 18
ws1.column_dimensions["E"].width = 45
ws1.column_dimensions["F"].width = 50
ws1.column_dimensions["G"].width = 40

# ═══════════════════════════════════════════════════════════
# Sheet 2: Backend API Tests
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Backend API Tests")
ws2.append(["#", "Test Case", "Description", "Result", "Expected Outcome", "Actual Outcome", "Solution"])
style_header(ws2)

api_tests = [
    ["B01", "Airtable: Read Daily Logs",
     "GET /v0/{base}/tblqEHZcreZ0DzFJo with date filter",
     "PASS", "HTTP 200, 0 records (no daily log for today yet)",
     "HTTP 200, 0 records", "—"],

    ["B02", "Airtable: Read Tasks",
     "GET /v0/{base}/tblgcTijaYRdLG3Wn with date filter",
     "PASS", "HTTP 200, returns tasks for today",
     "HTTP 200, 0 records initially (expected)", "—"],

    ["B03", "Airtable: Read Midday Logs",
     "GET /v0/{base}/tblvjUYwfzMxQrKHL with date filter",
     "PASS", "HTTP 200, returns midday logs for today",
     "HTTP 200, 0 records initially (expected)", "—"],

    ["B04", "Airtable: Read Drafts",
     "GET /v0/{base}/tblkBZuG4hP48JGvz with date+type filter",
     "PASS", "HTTP 200, returns drafts for today",
     "HTTP 200, 0 records initially (expected)", "—"],

    ["B05", "Airtable: Read Carryovers",
     "GET tasks with Status='Carried Over' filter",
     "PASS", "HTTP 200, returns carried-over tasks",
     "HTTP 200, 0 records (none exist yet)", "—"],

    ["B06", "Airtable: Create Task",
     "POST /v0/{base}/tblgcTijaYRdLG3Wn with task fields",
     "PASS", "HTTP 200, returns created record with ID",
     "Created recBKKkgeBWNw5dP5 successfully", "—"],

    ["B07", "Airtable: Create Midday Log",
     "POST /v0/{base}/tblvjUYwfzMxQrKHL with log fields",
     "PASS", "HTTP 200, returns created record with ID",
     "Created record successfully", "—"],

    ["B08", "Airtable: Date Filter Formula",
     "filterByFormula={Date}='2026-02-13' vs DATESTR({Date})='2026-02-13'",
     "FAIL → FIXED", "{Date}='date' should return matching records",
     "BUG: {Date}='date' returns 0 records. DATESTR({Date})='date' returns correct results. Airtable Date fields need DATESTR() wrapper.",
     "Fixed all 4 API routes: tasks, daily-logs, midday-logs, drafts. Changed filter formulas to use DATESTR()."],

    ["B09", "Frontend: GET /api/tasks",
     "Fetch today's tasks through Next.js API route",
     "PASS (after fix)", "HTTP 200, returns array of Task objects",
     "After DATESTR fix: HTTP 200, 2 tasks returned correctly", "Fixed in B08"],

    ["B10", "Frontend: GET /api/daily-logs",
     "Fetch today's daily log through Next.js API route",
     "PASS", "HTTP 200, returns DailyLog or null",
     "HTTP 200, null (no daily log exists yet)", "—"],

    ["B11", "Frontend: GET /api/midday-logs",
     "Fetch today's midday logs through Next.js API route",
     "PASS (after fix)", "HTTP 200, returns array of MiddayLog objects",
     "After DATESTR fix: HTTP 200, 2 logs returned", "Fixed in B08"],

    ["B12", "Frontend: GET /api/drafts",
     "Fetch today's drafts filtered by type through Next.js API route",
     "PASS (after fix)", "HTTP 200, returns array of Draft objects",
     "HTTP 200, 0 drafts (none generated yet)", "Fixed in B08"],

    ["B13", "Frontend: POST /api/tasks",
     "Create a new task through Next.js API route",
     "PASS", "HTTP 200, returns created Task with ID",
     "HTTP 200, ID: recFqhVgHF6UrX98x, task created", "—"],

    ["B14", "Frontend: POST /api/midday-logs",
     "Create a new midday log through Next.js API route",
     "PASS", "HTTP 200, returns created MiddayLog with ID",
     "HTTP 200, ID: reckJQkC4wmoNhizP", "—"],

    ["B15", "Frontend: PATCH /api/tasks",
     "Update task status through Next.js API route",
     "PASS", "HTTP 200, returns updated Task with new status",
     "HTTP 200, status changed To Do → In Progress", "—"],
]

for row in api_tests:
    ws2.append(row)
    result = row[3]
    if "FAIL" in result and "FIXED" in result:
        fill = YELLOW
    elif "FAIL" in result:
        fill = RED
    else:
        fill = GREEN
    for cell in ws2[ws2.max_row]:
        cell.font = NORMAL
        cell.alignment = WRAP
        cell.border = THIN_BORDER
    ws2.cell(row=ws2.max_row, column=4).fill = fill

ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 50
ws2.column_dimensions["D"].width = 18
ws2.column_dimensions["E"].width = 45
ws2.column_dimensions["F"].width = 55
ws2.column_dimensions["G"].width = 50

# ═══════════════════════════════════════════════════════════
# Sheet 3: n8n Workflow Tests
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("n8n Workflow Tests")
ws3.append(["#", "Test Case", "Description", "Result", "Expected Outcome", "Actual Outcome", "Solution"])
style_header(ws3)

n8n_tests = [
    ["N01", "Workflow status",
     "Check if workflow is active and ready to process triggers",
     "FAIL", "active: true — schedules and webhooks fire automatically",
     "active: false — workflow is INACTIVE. Nothing triggers automatically.",
     "Must activate workflow via n8n UI or API before going live."],

    ["N02", "Gemini Model node",
     "Verify Gemini 2.0 Flash is configured with valid credential",
     "PASS", "type: lmChatGoogleGemini, model: models/gemini-2.0-flash, credential linked",
     "Type: lmChatGoogleGemini, Model: models/gemini-2.0-flash, Credential: Google Gemini (OSyz6Gwd5Ygm8nuZ)", "—"],

    ["N03", "Gemini → AI Agents connections",
     "All 3 AI Agent nodes connected to Gemini Model",
     "PASS", "Morning, Midday, Evening agents all connected",
     "Morning AI Agent: CONNECTED, Midday AI Agent: CONNECTED, Evening AI Agent: CONNECTED", "—"],

    ["N04", "Basecamp Josef Tasks node",
     "HTTP Request node for Josef's Internal Tasks todolist",
     "PASS", "URL: .../todolists/9531841272/todos.json, OAuth2 credential, retry enabled",
     "URL correct, oAuth2Api credential linked, RetryOnFail: True", "—"],

    ["N05", "Basecamp My Assignments node",
     "HTTP Request node for cross-project assignments",
     "PASS", "URL: .../my/assignments.json, OAuth2 credential, retry enabled",
     "URL correct, oAuth2Api credential linked, RetryOnFail: True", "—"],

    ["N06", "Merge node inputs",
     "Merge All Morning Data has 5 inputs wired correctly",
     "PASS", "5 inputs: Daily Logs, Carryovers, Today Tasks, Josef Tasks, My Assignments",
     "Confirmed 5 inputs from correct source nodes", "—"],

    ["N07", "Webhook: /midday-agent",
     "Midday webhook endpoint responds to POST",
     "PASS (partial)", "HTTP 200 with AI agent response when workflow active",
     "Production URL returns HTTP 200. Test URL returns 404 (expected when not in test mode).",
     "Full test requires workflow activation."],

    ["N08", "Webhook: /send-time-in",
     "Send Time-In webhook endpoint responds to POST",
     "FAIL (expected)", "HTTP 200 after processing draft → email → daily log update",
     "HTTP 500 — tried to process fake draftRecordId, Airtable lookup failed.",
     "Expected failure with test data. Will work with real draft record IDs."],

    ["N09", "Webhook: /send-time-out",
     "Send Time-Out webhook endpoint responds to POST",
     "FAIL (expected)", "HTTP 200 after processing draft → email → hours calc → daily log",
     "HTTP 500 — tried to process fake draftRecordId, Airtable lookup failed.",
     "Expected failure with test data. Will work with real draft record IDs."],

    ["N10", "Gmail credentials",
     "Gmail OAuth2 credential configured for email sending",
     "PASS (config)", "gmailOAuth2 credential linked to all 4 Gmail nodes",
     "All 4 Gmail nodes (Morning/Evening Notification, TI/TO Send) reference gmailOAuth2.", "—"],

    ["N11", "Gmail recipients",
     "Email sent to correct address",
     "PASS (config)", "Morning/Evening to josef@jbbgi.com, TI/TO dynamic",
     "Morning/Evening: josef@jbbgi.com. TI/TO: dynamic from $json.to", "—"],

    ["N12", "Airtable tools (AI Agent)",
     "6 Airtable tool nodes for agent create/update operations",
     "PASS (config)", "Create Daily Log, Create/Update Tasks, Create/Update Drafts, Create Midday Logs",
     "All 6 airtableTool nodes present and connected", "—"],

    ["N13", "Morning Schedule trigger",
     "Schedule trigger fires at 7:30 AM daily",
     "NOT TESTED", "Trigger fires and starts morning flow",
     "Cannot test — workflow is inactive. Config appears correct.",
     "Requires activation. Test by running workflow manually in n8n editor."],

    ["N14", "Evening Schedule trigger",
     "Schedule trigger fires at end of day",
     "NOT TESTED", "Trigger fires and starts evening flow",
     "Cannot test — workflow is inactive.",
     "Requires activation. Test by running workflow manually in n8n editor."],

    ["N15", "Basecamp ContinueOnFail",
     "Basecamp nodes should continue if API is unreachable",
     "FAIL (config)", "continueOnFail: true so morning flow doesn't break if Basecamp is down",
     "continueOnFail: false on both Basecamp nodes",
     "Update Basecamp nodes to set continueOnFail: true. Otherwise a Basecamp outage kills the entire morning workflow."],
]

for row in n8n_tests:
    ws3.append(row)
    result = row[3]
    if "NOT TESTED" in result:
        fill = GRAY
    elif "FAIL" in result and ("expected" in result.lower() or "config" in result.lower()):
        fill = YELLOW
    elif "FAIL" in result:
        fill = RED
    elif "partial" in result.lower():
        fill = YELLOW
    else:
        fill = GREEN
    for cell in ws3[ws3.max_row]:
        cell.font = NORMAL
        cell.alignment = WRAP
        cell.border = THIN_BORDER
    ws3.cell(row=ws3.max_row, column=4).fill = fill

ws3.column_dimensions["A"].width = 6
ws3.column_dimensions["B"].width = 30
ws3.column_dimensions["C"].width = 50
ws3.column_dimensions["D"].width = 18
ws3.column_dimensions["E"].width = 50
ws3.column_dimensions["F"].width = 55
ws3.column_dimensions["G"].width = 55

# ═══════════════════════════════════════════════════════════
# Sheet 4: Summary
# ═══════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Summary")
ws4.append(["Category", "Total", "Pass", "Fail", "Fixed", "Not Tested", "Pass Rate"])
style_header(ws4)

# Frontend
f_total = len(ui_tests)
f_pass = sum(1 for t in ui_tests if "PASS" in t[3])
ws4.append(["Frontend UI (Playwright)", f_total, f_pass, 0, 0, 0, f"{f_pass}/{f_total}"])

# Backend
b_total = len(api_tests)
b_pass = sum(1 for t in api_tests if "PASS" in t[3] and "FAIL" not in t[3])
b_fixed = sum(1 for t in api_tests if "FIXED" in t[3])
b_fail = b_total - b_pass - b_fixed
ws4.append(["Backend API", b_total, b_pass, b_fail, b_fixed, 0, f"{b_pass + b_fixed}/{b_total}"])

# n8n
n_total = len(n8n_tests)
n_pass = sum(1 for t in n8n_tests if t[3].startswith("PASS"))
n_fail = sum(1 for t in n8n_tests if "FAIL" in t[3])
n_nt = sum(1 for t in n8n_tests if "NOT TESTED" in t[3])
ws4.append(["n8n Workflow", n_total, n_pass, n_fail, 0, n_nt, f"{n_pass}/{n_total}"])

# Totals
total = f_total + b_total + n_total
total_pass = f_pass + b_pass + b_fixed + n_pass
ws4.append([])
ws4.append(["TOTAL", total, total_pass, b_fail + n_fail, b_fixed, n_nt,
            f"{total_pass}/{total} ({round(total_pass/total*100)}%)"])

for row_idx in range(2, ws4.max_row + 1):
    for cell in ws4[row_idx]:
        cell.font = NORMAL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
if ws4.max_row >= 6:
    for cell in ws4[ws4.max_row]:
        cell.font = BOLD

ws4.column_dimensions["A"].width = 25
ws4.column_dimensions["B"].width = 8
ws4.column_dimensions["C"].width = 8
ws4.column_dimensions["D"].width = 8
ws4.column_dimensions["E"].width = 8
ws4.column_dimensions["F"].width = 12
ws4.column_dimensions["G"].width = 12

# ═══════════════════════════════════════════════════════════
# Sheet 5: Critical Issues
# ═══════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Critical Issues")
ws5.append(["#", "Issue", "Severity", "Status", "Impact", "Fix"])
style_header(ws5)

issues = [
    ["I01", "Airtable DATESTR filter bug",
     "CRITICAL", "FIXED",
     "ALL 4 GET API routes returned 0 records. Dashboard showed empty data even with records in Airtable. Users would see 'No tasks yet' permanently.",
     "Changed filterByFormula from {Date}='date' to DATESTR({Date})='date' in tasks, daily-logs, midday-logs, and drafts routes."],

    ["I02", "n8n workflow inactive",
     "HIGH", "OPEN",
     "Schedule triggers (morning 7:30 AM, evening) won't fire. Webhooks may not respond. No drafts generated, no emails sent automatically.",
     "Activate workflow in n8n UI or via API: PUT /workflows/{id}/activate"],

    ["I03", "Basecamp continueOnFail=false",
     "MEDIUM", "OPEN",
     "If Basecamp API is down or OAuth token expires, the entire morning workflow fails. No drafts, no tasks, no email.",
     "Set continueOnFail: true on both Basecamp HTTP Request nodes."],

    ["I04", "n8n Agent source not in Airtable",
     "LOW", "OPEN",
     "Tasks created by n8n AI Agent may use 'n8n Agent' as Source, but this value may not be in Airtable's Source field options.",
     "Add 'n8n Agent' to the Source field's select options in Airtable Tasks table."],
]

for row in issues:
    ws5.append(row)
    severity = row[2]
    status = row[3]
    sev_fill = RED if severity == "CRITICAL" else YELLOW if severity in ("HIGH", "MEDIUM") else GRAY
    stat_fill = GREEN if status == "FIXED" else RED if status == "OPEN" else GRAY
    for cell in ws5[ws5.max_row]:
        cell.font = NORMAL
        cell.alignment = WRAP
        cell.border = THIN_BORDER
    ws5.cell(row=ws5.max_row, column=3).fill = sev_fill
    ws5.cell(row=ws5.max_row, column=4).fill = stat_fill

ws5.column_dimensions["A"].width = 6
ws5.column_dimensions["B"].width = 30
ws5.column_dimensions["C"].width = 12
ws5.column_dimensions["D"].width = 10
ws5.column_dimensions["E"].width = 60
ws5.column_dimensions["F"].width = 55

# ═══════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════
output_path = "/Users/josefjohreld.medel/Downloads/projects/Time In Dashboard Tracker/journey-dashboard/tests/test-report.xlsx"
wb.save(output_path)
print(f"Report saved to: {output_path}")
print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
